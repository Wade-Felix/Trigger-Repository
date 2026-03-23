# -*- coding: utf-8 -*-
"""
定价预览表生成器

从飞书读取产品数据，按各 eBay 店铺策略计算最终售价，
输出 Excel 预览文件供人工审核。

输出列：MSKU | 中间定价 | nimo-official | BESTPTV | nimooutlet | nimodeals | nimo-direct
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path

from pricing.feishu_reader import FeishuProductRecord, read_feishu_products
from pricing.strategies import STRATEGY_REGISTRY

logger = logging.getLogger(__name__)

_STORE_COLUMNS = ["nimo-official", "BESTPTV", "nimooutlet", "nimodeals", "nimo-direct"]


def _build_rows(records: list[FeishuProductRecord]) -> list[dict]:
    rows = []
    for rec in records:
        row: dict = {"MSKU": rec.msku, "中间定价": rec.base_price}
        for store in _STORE_COLUMNS:
            strategy = STRATEGY_REGISTRY.get(store)
            if strategy is None:
                row[store] = ""
                continue
            try:
                row[store] = strategy.compute(rec.base_price)
            except (ValueError, TypeError) as exc:
                logger.warning("计算失败 msku=%s store=%s: %s", rec.msku, store, exc)
                row[store] = ""
        rows.append(row)
    return rows


def _write_xlsx(rows: list[dict], output_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "定价预览"

    headers = ["MSKU", "中间定价"] + _STORE_COLUMNS
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)


def _write_csv(rows: list[dict], output_path: Path) -> None:
    import csv
    headers = ["MSKU", "中间定价"] + _STORE_COLUMNS
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


async def generate_pricing_preview(output_dir: str | Path = ".") -> Path:
    """读取飞书数据，生成定价预览文件。

    优先输出 .xlsx，若 openpyxl 不可用则降级为 .csv。

    :param output_dir: 输出目录，默认为当前目录
    :returns: 生成的文件路径
    """
    logger.info("开始生成定价预览表……")

    records = await read_feishu_products()
    if not records:
        logger.warning("飞书未返回任何产品记录，预览表未生成。")
        raise RuntimeError("飞书未返回任何产品记录")

    rows = _build_rows(records)
    timestamp = datetime.now().strftime("%Y%m%d")
    output_dir = Path(output_dir)

    try:
        import openpyxl  # noqa: F401
        output_path = output_dir / f"定价预览_{timestamp}.xlsx"
        _write_xlsx(rows, output_path)
    except ImportError:
        output_path = output_dir / f"定价预览_{timestamp}.csv"
        _write_csv(rows, output_path)

    logger.info("定价预览表已生成：%s（共 %d 行）", output_path, len(rows))
    return output_path
