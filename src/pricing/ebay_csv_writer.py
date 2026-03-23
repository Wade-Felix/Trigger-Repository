# -*- coding: utf-8 -*-
"""
eBay 改价文件生成器

职责：读取 eBay 模板文件（单属性 / 多属性），
根据定价预览数据匹配更新价格，输出改价 CSV 文件。
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

_UNMATCHED_FILL = PatternFill("solid", fgColor="FFFF00")  # 黄色标注未匹配

# eBayUserID（模板中）→ 定价预览列名 的映射
_EBAY_USER_TO_STORE: dict[str, str] = {
    "NzuTUH3XQv-": "BESTPTV",
}


def _prepare_single_output(
    template_path: str | Path,
    output_dir: str | Path = ".",
) -> tuple[openpyxl.Workbook, Path]:
    """拷贝单属性模板，清空数据行，另存为新文件，只保留第一个 sheet 的表头。

    :param template_path: 单属性模板文件路径（.xlsx）
    :param output_dir: 输出目录，默认为当前目录
    :returns: (清空后的 Workbook, 另存路径)
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"单属性模板文件不存在：{template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    # 保留第一行（表头），删除其余所有数据行
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # 另存为新文件
    timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d")
    output_path = Path(output_dir) / f"单属性改价_{timestamp}.xlsx"
    wb.save(output_path)

    logger.info(
        "单属性模板已另存为：%s（表头列数：%d，数据行已清空）",
        output_path,
        ws.max_column,
    )
    return wb, output_path


def _match_single_from_preview(
    preview_path: str | Path,
    template_path: str | Path,
    output_wb: openpyxl.Workbook,
    output_path: str | Path,
) -> int:
    """循环定价预览表的每一行，按 MSKU 匹配单属性模板的 SKU 列，
    将匹配到的行拷贝至输出文件，并按 eBayUserID 对应店铺价格更新 StartPrice。

    :param preview_path: 定价预览 xlsx 路径
    :param template_path: 单属性模板 xlsx 路径（数据来源）
    :param output_wb: 已清空表头的输出 Workbook
    :param output_path: 输出文件保存路径
    :returns: 匹配写入的行数
    """
    preview_wb = openpyxl.load_workbook(preview_path)
    preview_ws = preview_wb.worksheets[0]

    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.worksheets[0]
    output_ws = output_wb.worksheets[0]

    # 解析定价预览表头，建立店铺名 → 列索引（0-based）映射
    preview_header = [cell.value for cell in preview_ws[1]]
    store_col_map: dict[str, int] = {
        col: idx for idx, col in enumerate(preview_header) if idx >= 2
    }

    # 构建预览表 MSKU → 该行数据 映射
    msku_to_preview_row: dict[str, tuple] = {}
    for row in preview_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            msku_to_preview_row[str(row[0]).strip()] = row

    # 解析单属性模板表头
    template_header = [cell.value for cell in template_ws[1]]
    sku_col_idx = template_header.index("SKU")
    platform_sku_col_idx = template_header.index("PlatformSKU") if "PlatformSKU" in template_header else None
    ebay_user_col_idx = template_header.index("eBayUserID")
    start_price_col_idx = template_header.index("StartPrice")

    # 准备未匹配输出文件
    timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d")
    unmatched_path = Path(output_path).parent / f"单属性未匹配_{timestamp}.xlsx"
    unmatched_wb = openpyxl.load_workbook(template_path)
    unmatched_ws = unmatched_wb.worksheets[0]
    if unmatched_ws.max_row > 1:
        unmatched_ws.delete_rows(2, unmatched_ws.max_row - 1)

    matched = 0
    unmatched = 0
    for row in template_ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        template_row = list(row)

        # SKU 优先，为空时降级用 PlatformSKU 匹配
        sku = row[sku_col_idx]
        if sku:
            match_key = str(sku).strip()
            highlight_col = sku_col_idx + 1
        elif platform_sku_col_idx is not None and row[platform_sku_col_idx]:
            match_key = str(row[platform_sku_col_idx]).strip()
            highlight_col = platform_sku_col_idx + 1
        else:
            # SKU 和 PlatformSKU 均为空，归入未匹配
            unmatched_ws.append(template_row)
            unmatched += 1
            continue

        if match_key not in msku_to_preview_row:
            unmatched_ws.append(template_row)
            unmatched_ws.cell(unmatched_ws.max_row, highlight_col).fill = _UNMATCHED_FILL
            unmatched += 1
            continue

        store_name = str(template_row[ebay_user_col_idx]).strip() if template_row[ebay_user_col_idx] else ""
        store_name = _EBAY_USER_TO_STORE.get(store_name, store_name)
        if store_name in store_col_map:
            price = msku_to_preview_row[match_key][store_col_map[store_name]]
            if price:
                template_row[start_price_col_idx] = round(float(price), 2)
        else:
            logger.warning("店铺 %s 在定价预览中无对应价格列，match_key=%s", store_name, match_key)

        output_ws.append(template_row)
        matched += 1

    output_wb.save(output_path)
    unmatched_wb.save(unmatched_path)
    logger.info("单属性匹配完成：匹配 %d 行，未匹配 %d 行，已保存至 %s / %s",
                matched, unmatched, output_path, unmatched_path)
    return matched


def _prepare_multi_output(
    template_path: str | Path,
    output_dir: str | Path = ".",
) -> tuple[openpyxl.Workbook, Path]:
    """拷贝多属性模板，清空数据行，另存为新文件。"""
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"多属性模板文件不存在：{template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d")
    output_path = Path(output_dir) / f"多属性改价_{timestamp}.xlsx"
    wb.save(output_path)

    logger.info(
        "多属性模板已另存为：%s（表头列数：%d，数据行已清空）",
        output_path,
        ws.max_column,
    )
    return wb, output_path


def _match_multi_from_preview(
    preview_path: str | Path,
    template_path: str | Path,
    output_wb: openpyxl.Workbook,
    output_path: str | Path,
) -> int:
    """按 listing 块匹配多属性模板，将命中的整块拷贝至输出并更新 V_Price。

    匹配规则：listing 块内任意 V_SKU 存在于预览表 MSKU → 整块写入输出。
    拷贝时：匹配到的变体行更新 V_Price，未匹配的保留原价。
    去重：按 eBayItemID 只写入一次。
    """
    preview_wb = openpyxl.load_workbook(preview_path)
    preview_ws = preview_wb.worksheets[0]

    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.worksheets[0]
    output_ws = output_wb.worksheets[0]

    # 预览表：店铺列映射 & MSKU → 行数据
    preview_header = [cell.value for cell in preview_ws[1]]
    store_col_map: dict[str, int] = {
        col: idx for idx, col in enumerate(preview_header) if idx >= 2
    }
    msku_to_preview_row: dict[str, tuple] = {}
    for row in preview_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            msku_to_preview_row[str(row[0]).strip()] = row

    # 多属性模板列索引
    template_header = [cell.value for cell in template_ws[1]]
    ebay_user_col_idx = template_header.index("eBayUserID")
    ebay_item_col_idx = template_header.index("eBayItemID")
    vsku_col_idx = template_header.index("V_SKU")
    vprice_col_idx = template_header.index("V_Price")

    # 将模板所有数据行分组为 listing 块
    # 每块：{"store": str, "item_id": str, "rows": [list, ...]}
    blocks: list[dict] = []
    current_block: dict | None = None

    for row in template_ws.iter_rows(min_row=2, values_only=True):
        ebay_user = row[ebay_user_col_idx]
        if ebay_user:  # 父行 → 新 listing 块
            current_block = {
                "store": _EBAY_USER_TO_STORE.get(str(ebay_user).strip(), str(ebay_user).strip()),
                "item_id": str(row[ebay_item_col_idx]).strip() if row[ebay_item_col_idx] else "",
                "rows": [list(row)],
            }
            blocks.append(current_block)
        elif current_block is not None:  # 子行 → 追加到当前块
            current_block["rows"].append(list(row))

    # 准备未匹配输出文件
    timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d")
    unmatched_path = Path(output_path).parent / f"多属性未匹配_{timestamp}.xlsx"
    unmatched_wb = openpyxl.load_workbook(template_path)
    unmatched_ws = unmatched_wb.worksheets[0]
    if unmatched_ws.max_row > 1:
        unmatched_ws.delete_rows(2, unmatched_ws.max_row - 1)

    seen_item_ids: set[str] = set()
    matched_blocks = 0
    unmatched_blocks = 0

    for block in blocks:
        item_id = block["item_id"]
        if item_id in seen_item_ids:
            continue
        seen_item_ids.add(item_id)

        store = block["store"]
        store_price_col = store_col_map.get(store)

        hit = any(
            str(r[vsku_col_idx]).strip() in msku_to_preview_row
            for r in block["rows"]
            if r[vsku_col_idx]
        )

        if not hit:
            for row in block["rows"]:
                unmatched_ws.append(row)
                vsku = str(row[vsku_col_idx]).strip() if row[vsku_col_idx] else ""
                if vsku:
                    unmatched_ws.cell(unmatched_ws.max_row, vsku_col_idx + 1).fill = _UNMATCHED_FILL
            unmatched_blocks += 1
            continue

        for row in block["rows"]:
            vsku = str(row[vsku_col_idx]).strip() if row[vsku_col_idx] else ""
            if vsku and vsku in msku_to_preview_row and store_price_col is not None:
                price = msku_to_preview_row[vsku][store_price_col]
                if price:
                    row[vprice_col_idx] = round(float(price), 2)
            output_ws.append(row)

        matched_blocks += 1

    output_wb.save(output_path)
    unmatched_wb.save(unmatched_path)
    logger.info("多属性匹配完成：匹配 %d 块，未匹配 %d 块，已保存至 %s / %s",
                matched_blocks, unmatched_blocks, output_path, unmatched_path)
    return matched_blocks
